#NOTE: code by visual studio
#thư viện đọc Excel
from openpyxl import load_workbook
import pandas as pd
#thư viện xử lý mảng
import numpy as np
#thư viện vẽ đồ thị
from matplotlib import pyplot
import matplotlib.pyplot as plt

#đọc file Excel
wb = load_workbook(filename='So lieu du toan NSNN nam 2019.xlsx', read_only=True)
ws = wb['B18']  #lấy dữ liệu ở sheet B18

# Lấy dữ liệu từ B13 -> C52, ở sheet B18
data_rows = []
for row in ws['B13':'C52']:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)
df = pd.DataFrame(data_rows)

# hàm sắp xếp, đồng thời đổi luôn vị trí của tên các cơ quan
def sapXep(a):
    #Thuật toán sắp xếp chọn (Selection Sort)
    for i in range(0,len(a) - 1):
        min = i
        for j in range(i,len(a)):
            if (a[j][1] < a[min][1]):
                min = j
        #sắp xếp theo số tiền & đồng thời thay đổi vị trí cơ quan tương ứng
        temp = a[min][1]
        temp_name = a[min][0]
        a[min][1] = a[i][1]
        a[min][0] = a[i][0]
        a[i][1] = temp
        a[i][0] = temp_name

    return a

""" Xây dựng hàm tính phần trăm (%). 
    Ta tạo một mảng mới để lưu các các phần tử phần trăm"""
def tinhPhanTram(mang):
    #tính tổng số tiền
    sum = 0
    for i in range(len(mang)):
        sum = sum + mang[i][1]
    mangPhanTram = np.zeros(len(mang)) # khởi tạo mảng
    for i in range(len(mang)):
        mangPhanTram[i] = (mang[i][1] / sum) *100  # tính phần trăm (%) từng phần tử
    return mangPhanTram

# Hàm thực hiện YÊU CẦU 1
def yeuCau1(mang, mangPhanTram, k):
    for i in range(len(mang) - 1, len(mang) - k - 1, -1): #Lấy từ phần tử cuối (gs: n) tới (n-k+1)
        print(mang[i][0], '---' , mangPhanTram[i], "%")

# Hàm thực hiện YÊU CẦU 2
"""đầu vào: mảng phần trăm (mangPhanTram) đã được sắp xếp
    và mảng để in ra tên cơ quan và số tiền
 (do mảng phần trăm được tính theo số tiền từ nhỏ đến lớn 
 nên mảng phần trăm cũng đã được sắp xếp tăng dần)"""
def yeuCau2(mang, mangPhanTram):
    sum = 0
    soCoQuan = 0
    #tìm số cơ quan
    for i in range(len(mangPhanTram) - 1, -1, -1): #duyệt tất cả phần tử của mảng
        soCoQuan = soCoQuan + 1
        sum = sum + mangPhanTram[i]
        if(sum >= 50):
            break
    #in ra các cơ quan đó
    print("Số cơ quan (k nhỏ nhất có thể) mà tổng số tiền chi cho các lĩnh vực này chiếm tới 50% là "\
        ,soCoQuan, "Gồm có: ")
    for i in range(soCoQuan):
        print(mang[(len(mangPhanTram) - 1 - i)][0], 'Số tiền : ' , mang[(len(mangPhanTram) - 1 - i)][1])
    return soCoQuan

# Hàm thực hiện YÊU CẦU 3
"""sử dụng thuật toán tìm kiếm tuần tự"""
def yeuCau3(npData):
    for i in range(len(npData)):   
        if(npData[i][0] == "Bộ Giáo dục và Đào tạo"):
            print("Số tiền Bộ Giáo dục và Đào tạo nhận được là: ", npData[i][1], "Triệu đồng")

# Hàm thực hiện YÊU CẦU 4
def yeuCau4(npData):
    mangVe = []
    for i in range(len(npData) - 1, len(npData) - 11, -1):
        mangVe.append(npData[i][1])
    fig, ax = plt.subplots()
    ax.bar(range(1,11),mangVe, align='center')
    ax.set_xticks(np.arange(11)) # danh sách các cột được gán nhãn trên trục x
    ax.set_xlabel('10 cơ quan')
    ax.set_ylabel('Số tiền')
    ax.set_title('Biểu đồ biểu diễn số tiền mà 10 cơ quan được nhận nhiều nhất')
    plt.show()

# Hàm thực hiện YÊU CẦU 5
def yeuCau5(npData):
    mangVe = []
    for i in range(len(npData)):
        mangVe.append(npData[i][1])
    fig, ax = plt.subplots()
    ax.set_xlabel('Các cơ quan')
    ax.set_ylabel('Số tiền')
    ax.set_title('Đồ thị biểu diễn số tiền mà tất cả các cơ quan nhận được')
    plt.plot(range(1,len(npData) + 1), mangVe)
    plt.show()

# Hàm thực hiện YÊU CẦU 6
""" Chỉ cần in lại mảng theo chiều ngược lại """
def yeuCau6(npData) :
    for i in range(len(npData) - 1, -1, -1):
        print ("Tên cơ quan: ", npData[i][0], "---", "Số tiền: ", npData[i][1])








# chuyển dữ liệu sang kiểu numpy
npData = np.array(df)
npData = sapXep(npData)

mangPhanTram = tinhPhanTram(npData)

##Truy xuất YÊU CẦU 1
#k = int (input("Nhap k: "))
#if(k > len(mangPhanTram)):
#    print("So luong lon hon so phan tu")
#else:
#    print(k ,"cơ quan dự kiến nhận được số tiền nhiều nhất là: ")
#    yeuCau1(npData,mangPhanTram,k)

##truy xuất yêu cầu 2
#yeuCau2(npData, mangPhanTram)


##Truy xuất YÊU CẦU 3
#yeuCau3(npData)

##Truy xuất YÊU CẦU 4
#yeuCau4(npData)

##Truy xuất YÊU CẦU 5
#yeuCau5(npData)

##Truy xuất YÊU CẦU 6
#print("Các cơ quan theo thứ tự giảm dần của số tiền nhận được từ ngân sách là: ")
#yeuCau6(npData)